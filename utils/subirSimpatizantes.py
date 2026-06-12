"""
subirSimpatizantes.py — Carga masiva de simpatizantes desde Excel a la BD
Ubicación: utils/subirSimpatizantes.py

Uso:
    python utils/subirSimpatizantes.py

Requisitos: openpyxl, python-dotenv (ya en tu venv)
"""

import os
import sys

# ── Bootstrap Django ──────────────────────────────────────────────────────────
RAIZ = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, RAIZ)

from dotenv import load_dotenv
load_dotenv(os.path.join(RAIZ, ".env"))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "app.settings.development")

import logging
logging.disable(logging.CRITICAL)

import django
try:
    django.setup()
except Exception as e:
    logging.disable(logging.NOTSET)
    print(f"\n  ✗  Error iniciando Django: {e}")
    print("     Verifica que el .env tenga PGDATABASE, PGUSER, PGPASSWORD, PGHOST.")
    input("\n  Presiona Enter para cerrar...")
    sys.exit(1)

logging.disable(logging.NOTSET)

from django.db import connection, DatabaseError
from app import db

# ── Colores consola ───────────────────────────────────────────────────────────
os.system("color")

class C:
    RESET  = "\033[0m"
    BOLD   = "\033[1m"
    GREEN  = "\033[92m"
    YELLOW = "\033[93m"
    RED    = "\033[91m"
    CYAN   = "\033[96m"
    DIM    = "\033[2m"

def ok(msg):   print(f"  {C.GREEN}✓{C.RESET}  {msg}")
def info(msg): print(f"  {C.CYAN}i{C.RESET}  {msg}")
def warn(msg): print(f"  {C.YELLOW}!{C.RESET}  {msg}")
def err(msg):  print(f"  {C.RED}✗{C.RESET}  {msg}")

def header(msg):
    print(f"\n{C.BOLD}{C.CYAN}{'─' * 56}{C.RESET}")
    print(f"{C.BOLD}{C.CYAN}  {msg}{C.RESET}")
    print(f"{C.BOLD}{C.CYAN}{'─' * 56}{C.RESET}")

def separador():
    print(f"  {C.DIM}{'·' * 52}{C.RESET}")

# ── Columnas esperadas en el Excel ────────────────────────────────────────────
# (nombre_columna, obligatorio, tipo)
COLUMNAS = [
    ("nombre",           True,  str),
    ("cedula",           True,  str),
    ("telefono",         False, str),
    ("edad",             True,  int),
    ("ocupacion",        True,  str),
    ("lugar_votacion",   True,  str),
    ("puesto_votacion",  True,  str),
    ("mesa_votacion",    True,  str),
    ("opinion_politica", False, str),
    ("barrio_nombre",    False, str),
]

# ── Helpers BD ────────────────────────────────────────────────────────────────

def verificar_conexion():
    with connection.cursor() as cur:
        cur.execute("SELECT 1")

def obtener_barrios() -> dict:
    with connection.cursor() as cur:
        cur.execute(f"SELECT id, nombre FROM {db.barrio}")
        return {row[1].strip().lower(): str(row[0]) for row in cur.fetchall()}

def cedula_existe(cedula: str) -> bool:
    with connection.cursor() as cur:
        cur.execute(
            f"SELECT 1 FROM {db.simpatizante} WHERE cedula = %s",
            [cedula],
        )
        return cur.fetchone() is not None

def insertar_simpatizante(payload: dict) -> str:
    with connection.cursor() as cur:
        cur.execute(
            f"""INSERT INTO {db.simpatizante}
                (nombre, cedula, telefono, edad, ocupacion,
                 lugar_votacion, puesto_votacion, mesa_votacion,
                 opinion_politica, barrio_id)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            [
                payload["nombre"],
                payload["cedula"],
                payload["telefono"],
                payload["edad"],
                payload["ocupacion"],
                payload["lugar_votacion"],
                payload["puesto_votacion"],
                payload["mesa_votacion"],
                payload["opinion_politica"],
                payload["barrio_id"],
            ],
        )
        return str(cur.fetchone()[0])

# ── Lectura y validación del Excel ────────────────────────────────────────────

def leer_excel(ruta: str):
    import openpyxl

    ruta = ruta.strip().strip('"').strip("'")

    if not os.path.exists(ruta):
        err(f"Archivo no encontrado: {ruta}")
        return None, None

    if not ruta.lower().endswith((".xlsx", ".xls")):
        warn("El archivo no tiene extensión .xlsx — se intentará abrir de todos modos.")

    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
    except Exception as e:
        err(f"No se pudo abrir el archivo: {e}")
        return None, None

    ws = wb.active

    if ws.max_row < 2:
        warn("El archivo está vacío o solo tiene encabezados.")
        return [], []

    # Encabezados fila 1
    encabezados = [
        str(ws.cell(row=1, column=c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
    ]

    col_esperadas = [col for col, _, _ in COLUMNAS]
    faltantes     = [c for c in col_esperadas if c not in encabezados]
    if faltantes:
        err("Columnas faltantes en el archivo:")
        for f in faltantes:
            print(f"       - {f}")
        info("Las columnas deben llamarse exactamente así (minúsculas):")
        print(f"       {', '.join(col_esperadas)}")
        return None, None

    idx = {col: encabezados.index(col) for col in col_esperadas}

    filas_validas      = []
    errores_validacion = []

    for fila_num in range(2, ws.max_row + 1):
        valores = [ws.cell(row=fila_num, column=c + 1).value for c in range(ws.max_column)]

        if all(v is None or str(v).strip() == "" for v in valores):
            continue

        fila         = {}
        errores_fila = []

        for col, obligatorio, tipo in COLUMNAS:
            raw = valores[idx[col]]
            val = str(raw).strip() if raw is not None else ""

            if obligatorio and val == "":
                errores_fila.append(f"'{col}' es obligatorio y está vacío")
                continue

            if val == "":
                fila[col] = None
                continue

            if tipo == int:
                try:
                    convertido = int(float(val))
                    if convertido <= 0:
                        errores_fila.append(f"'{col}' debe ser entero positivo (valor: {val})")
                        continue
                    fila[col] = convertido
                except (ValueError, TypeError):
                    errores_fila.append(f"'{col}' debe ser número entero (valor: {val})")
                    continue
            else:
                fila[col] = val

        if errores_fila:
            errores_validacion.append((fila_num, errores_fila))
        else:
            filas_validas.append(fila)

    return filas_validas, errores_validacion

# ── Flujo por archivo ─────────────────────────────────────────────────────────

def procesar_archivo():
    print()
    ruta = input("  Ruta del archivo Excel: ").strip()
    if not ruta:
        warn("No ingresaste ninguna ruta.")
        return

    header("Leyendo archivo")
    info(f"Archivo: {ruta}")

    filas, errores_val = leer_excel(ruta)

    if filas is None:
        return

    if errores_val:
        print()
        warn(f"Problemas encontrados en {len(errores_val)} fila(s):")
        for fila_num, errs in errores_val:
            for e in errs:
                print(f"     {C.YELLOW}Fila {fila_num}:{C.RESET} {e}")

        if not filas:
            err("No hay filas válidas para procesar.")
            return

        print()
        resp = input(
            f"  ¿Continuar solo con las {len(filas)} fila(s) válidas? [s/N]: "
        ).strip().lower()
        if resp != "s":
            info("Operación cancelada.")
            return

    if not filas:
        warn("No se encontraron filas con datos.")
        return

    ok(f"{len(filas)} fila(s) válida(s) listas para subir")

    # Verificar conexión
    header("Verificando conexión a la BD")
    try:
        verificar_conexion()
        ok("Conexión a PostgreSQL establecida")
    except DatabaseError as e:
        err(f"No se pudo conectar a la BD: {e}")
        info("Verifica que el .env tenga las variables PG* correctas y que la BD esté activa.")
        return

    # Cargar barrios
    header("Cargando catálogo de barrios")
    try:
        barrios = obtener_barrios()
        ok(f"{len(barrios)} barrio(s) registrado(s) en la BD")
    except DatabaseError as e:
        err(f"Error consultando barrios: {e}")
        return

    # Previsualización
    header("Previsualización")
    print()
    print(
        f"  {C.BOLD}"
        f"{'#':<4} {'Nombre':<28} {'Cédula':<14} "
        f"{'Edad':<5} {'Ocupación':<18} {'Barrio'}"
        f"{C.RESET}"
    )
    separador()
    for i, f in enumerate(filas[:10], 1):
        barrio_disp = f.get("barrio_nombre") or "—"
        print(
            f"  {i:<4}"
            f"{str(f.get('nombre',    ''))[:27]:<28}"
            f"{str(f.get('cedula',    '')):<14}"
            f"{str(f.get('edad',      '')):<5}"
            f"{str(f.get('ocupacion', ''))[:17]:<18}"
            f"{barrio_disp}"
        )
    if len(filas) > 10:
        info(f"... y {len(filas) - 10} fila(s) más no mostradas")

    print()
    resp = input(
        f"  ¿Subir {len(filas)} simpatizante(s) a la BD? [s/N]: "
    ).strip().lower()
    if resp != "s":
        info("Operación cancelada.")
        return

    # Inserción
    header("Insertando registros")
    print()

    insertados = 0
    duplicados = []
    sin_barrio = []
    errores_bd = []

    for i, fila in enumerate(filas, 1):
        cedula = fila["cedula"]
        nombre = fila["nombre"]
        pref   = f"  [{i}/{len(filas)}]"

        # Resolver barrio
        barrio_nombre = fila.get("barrio_nombre")
        barrio_id     = None
        if barrio_nombre:
            barrio_id = barrios.get(barrio_nombre.strip().lower())
            if not barrio_id:
                sin_barrio.append({"cedula": cedula, "nombre": nombre, "barrio": barrio_nombre})
                print(
                    f"{pref} {C.YELLOW}!{C.RESET}  "
                    f"Barrio '{barrio_nombre}' no existe — {nombre} se inserta sin barrio"
                )

        # Verificar duplicado
        try:
            if cedula_existe(cedula):
                duplicados.append({"cedula": cedula, "nombre": nombre})
                print(
                    f"{pref} {C.YELLOW}!{C.RESET}  "
                    f"Duplicado omitido — cédula {cedula} ya existe ({nombre})"
                )
                continue
        except DatabaseError as e:
            errores_bd.append({"cedula": cedula, "nombre": nombre, "error": str(e)})
            print(f"{pref} {C.RED}✗{C.RESET}  Error verificando cédula {cedula}: {e}")
            continue

        payload = {
            "nombre":           nombre,
            "cedula":           cedula,
            "telefono":         fila.get("telefono"),
            "edad":             fila["edad"],
            "ocupacion":        fila["ocupacion"],
            "lugar_votacion":   fila["lugar_votacion"],
            "puesto_votacion":  fila["puesto_votacion"],
            "mesa_votacion":    fila["mesa_votacion"],
            "opinion_politica": fila.get("opinion_politica"),
            "barrio_id":        barrio_id,
        }

        try:
            uuid = insertar_simpatizante(payload)
            print(
                f"{pref} {C.GREEN}✓{C.RESET}  "
                f"{nombre} ({cedula}) → id: {uuid[:8]}…"
            )
            insertados += 1
        except DatabaseError as e:
            errores_bd.append({"cedula": cedula, "nombre": nombre, "error": str(e)})
            print(f"{pref} {C.RED}✗{C.RESET}  Error insertando {nombre}: {e}")

    # Resumen
    header("Resumen de la carga")
    print()
    ok(f"Insertados correctamente : {insertados}")
    if duplicados:
        warn(f"Omitidos por duplicado   : {len(duplicados)}")
    if sin_barrio:
        warn(f"Insertados sin barrio    : {len(sin_barrio)}")
    if errores_bd:
        err(f"Errores de BD            : {len(errores_bd)}")

    if duplicados:
        print()
        info("Cédulas duplicadas omitidas:")
        for d in duplicados:
            print(f"       {d['cedula']} — {d['nombre']}")

    if sin_barrio:
        print()
        info("Barrios no encontrados (insertados sin barrio):")
        for s in sin_barrio:
            print(f"       '{s['barrio']}' — {s['nombre']} ({s['cedula']})")

    if errores_bd:
        print()
        info("Detalle de errores de BD:")
        for e in errores_bd:
            print(f"       {e['cedula']} — {e['nombre']}: {C.RED}{e['error']}{C.RESET}")

# ── Bucle principal ───────────────────────────────────────────────────────────

def main():
    header("PactoHistórico · Carga masiva de Simpatizantes")
    print()
    info("Columnas esperadas en el Excel:")
    print()
    for col, ob, _ in COLUMNAS:
        marca = f"{C.GREEN}✓ obligatorio{C.RESET}" if ob else f"{C.DIM}  opcional  {C.RESET}"
        print(f"       {col:<22} {marca}")

    while True:
        try:
            procesar_archivo()
        except KeyboardInterrupt:
            print()
            info("Interrumpido — volviendo al menú.")
        except Exception as e:
            err(f"Error inesperado: {e}")
            info("El programa sigue activo.")

        print()
        separador()
        print()
        print(f"  {C.BOLD}¿Qué deseas hacer?{C.RESET}")
        print(f"    {C.GREEN}[1]{C.RESET}  Subir otro archivo")
        print(f"    {C.RED}[2]{C.RESET}  Salir")
        print()

        while True:
            try:
                opcion = input("  Opción [1/2]: ").strip()
            except KeyboardInterrupt:
                opcion = "2"
            if opcion in ("1", "2"):
                break
            warn("Escribe 1 o 2.")

        if opcion == "2":
            print()
            ok("Hasta luego.")
            print()
            break

# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print()
        ok("Programa cerrado.")